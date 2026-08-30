"""项目包服务与结果 Excel 导出。

依据架构宪法 §10/§12 与 domain-model §快照、任务和结果/§对象生命周期 及 contracts §公共文件契约：

- export_package: 仅所有者；版本化清单(格式版本/清单/对象清单)，流式导出
  模型/配置/版本/数据集版本与溯源/历史结果证据与评估引用/内容校验；
  不含账号/权限/会话/全局配置/密钥（domain-model §对象生命周期、架构宪法 §16）；
- import_proposal: 导入前校验(格式/兼容性/清单/完整性，sha256 逐对象校验)；
  暂存对象 + 拟创建项目快照 + 分区提交内容 + 校验报告；
- confirm_import: 提交导入 — 每次导入创建新项目身份(不覆盖已有)，导入者成为
  所有者，原授权关系不迁移，历史结果作为证据来源保留(不伪造本地任务)；
- export_excel: 固定模板(标题中英双语，默认中文)，固定引用证据包与评估，
  不重新求解；查看者可导出 Excel，仅所有者可导出项目包；
- 下载授权: 短期单对象授权(HMAC 签名 token 含 object_id + 过期，过期 5 分钟)。

本层服务不主动 commit，事务边界由 API 层控制。
"""

from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import zipfile
from datetime import UTC, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from iesplan import __version__
from iesplan.config import settings
from iesplan.core.diagnostics import SEVERITY_ERROR
from iesplan.core.errors import AppError, ConflictError, ForbiddenError, NotFoundError
from iesplan.core.idgen import sha256_hex
from iesplan.core.jsonutil import jsonable
from iesplan.models.audit import ImportProposal
from iesplan.models.calc import CalcSnapshot, Task
from iesplan.models.dataset import Dataset, DatasetFile, DatasetVersion
from iesplan.models.identity import User
from iesplan.models.project import Draft, Project, ProjectVersion, VersionRef
from iesplan.models.result import EvidencePackage, ResultAssessment, ResultIndex
from iesplan.services import audit as audit_service
from iesplan.services import project as project_service
from iesplan.storage import add_ref, get_object, object_by_sha256, object_info, put_object

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

#: 项目包格式版本(主版本号兼容判定: 只接受 1.x)
PACKAGE_FORMAT_VERSION = "1.0"
#: 项目包媒体类型
PACKAGE_MEDIA_TYPE = "application/zip"
#: Excel 报告媒体类型
EXCEL_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
#: 下载授权有效期(秒, 短期单对象授权, 默认 5 分钟)
DOWNLOAD_TOKEN_TTL_SECONDS = 300
#: 包内禁止出现的清单键(账号/权限/会话/全局配置/密钥不得随包导出，见 domain-model §对象生命周期)
FORBIDDEN_MANIFEST_KEYS: frozenset[str] = frozenset(
    {"accounts", "permissions", "sessions", "global_config", "secrets"}
)
#: 项目包必需文件
REQUIRED_PACKAGE_FILES: tuple[str, ...] = ("manifest.json", "project.json", "draft.json")
#: 数据集版本内容媒体类型映射(format 校验)
_FORMAT_BY_MEDIA: dict[str, str] = {
    "text/csv; charset=utf-8": "csv",
    "text/csv": "csv",
    "application/json": "json",
}

# ---------------------------------------------------------------------------
# 异常
# ---------------------------------------------------------------------------


class ImportValidationError(AppError):
    """项目包校验失败(格式/兼容性/清单/完整性, HTTP 400)。

    携带 params.reasons 逐项说明失败原因; 校验失败即拒绝, 不创建任何记录。
    """

    code = "PKG-IMP-001"
    http_status = 400
    severity = SEVERITY_ERROR
    message_key = "ies.diag.pkg.invalid"

    def __init__(self, reasons: list[str], message: str = "") -> None:
        self.reasons = list(reasons)
        super().__init__(message or f"项目包校验失败: {'; '.join(reasons)}")


class DownloadTokenError(AppError):
    """下载授权无效/过期(HTTP 400)。"""

    code = "EXPORT-TOKEN-001"
    http_status = 400
    severity = SEVERITY_ERROR
    message_key = "ies.diag.export.token_invalid"


class PackageSizeError(AppError):
    """项目包超限(上传字节/条目数/单条目解压大小/总解压大小, HTTP 413)。

    在任何解压读取之前完成门禁, 防止 ZIP Bomb 消耗内存与 CPU。
    """

    code = "PKG-SIZE-001"
    http_status = 413
    severity = SEVERITY_ERROR
    message_key = "ies.diag.pkg.too_large"


# ---------------------------------------------------------------------------
# 下载授权: 短期单对象签名 token(绑定项目与签发用户 + 过期, 过期 5 分钟)
# ---------------------------------------------------------------------------

#: 项目包上传字节上限(压缩后; 与 Nginx client_max_body_size 对齐为 2GB)
MAX_PACKAGE_BYTES: int = 2 * 1024 * 1024 * 1024
#: 项目包 zip 最大条目数(含目录; 超限拒绝, 防数十万小文件)
MAX_PACKAGE_ENTRIES: int = 5000
#: 单条目解压后大小上限(512MB, 防单文件巨大膨胀)
MAX_PACKAGE_ENTRY_BYTES: int = 512 * 1024 * 1024
#: 全部条目解压后总大小上限(4GB, 防整体 ZIP Bomb)
MAX_PACKAGE_TOTAL_BYTES: int = 4 * 1024 * 1024 * 1024


def _token_sign(payload: str) -> str:
    """HMAC-SHA256 签名(密钥取 settings.secret_key)。"""
    return hmac.new(
        settings.secret_key.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256
    ).hexdigest()


def create_download_token(
    object_id: int,
    kind: str,
    *,
    project_id: int,
    user_id: int,
    ttl_seconds: int = DOWNLOAD_TOKEN_TTL_SECONDS,
) -> str:
    """签发短期单对象下载授权 token(绑定项目与签发用户，架构宪法 §16)。

    token = base64url(payload) + "." + hmac 签名; payload 含 object_id/kind/
    project_id/user_id/exp, 缺省 5 分钟过期。下载时必须验证 token 绑定的项目
    与当前会话用户一致, 防止跨项目对象下载。
    """
    exp = int(datetime.now(UTC).timestamp()) + max(int(ttl_seconds), 1)
    payload = base64.urlsafe_b64encode(
        json.dumps(
            {
                "object_id": int(object_id),
                "kind": str(kind),
                "project_id": int(project_id),
                "user_id": int(user_id),
                "exp": exp,
            },
            sort_keys=True,
        ).encode("utf-8")
    ).decode("ascii")
    return f"{payload}.{_token_sign(payload)}"


def verify_download_token(token: str, *, expected_kind: str | None = None) -> dict[str, Any]:
    """校验下载授权 token: 格式/签名/过期, 返回 {object_id, kind, project_id, user_id}。

    签名比较使用 hmac.compare_digest(常量时间, 防时序侧信道);
    签名不符/格式非法/已过期一律抛 DownloadTokenError。
    """
    if not isinstance(token, str) or "." not in token:
        raise DownloadTokenError("", params={"reason": "bad_token"})
    payload, sig = token.rsplit(".", 1)
    if not hmac.compare_digest(_token_sign(payload), sig):
        raise DownloadTokenError("", params={"reason": "bad_signature"})
    try:
        data = json.loads(base64.urlsafe_b64decode(payload.encode("ascii")))
    except (ValueError, UnicodeDecodeError) as exc:
        raise DownloadTokenError("", params={"reason": "bad_payload"}) from exc
    if not isinstance(data, dict) or "object_id" not in data or "exp" not in data:
        raise DownloadTokenError("", params={"reason": "bad_payload"})
    if int(data["exp"]) < int(datetime.now(UTC).timestamp()):
        raise DownloadTokenError("", params={"reason": "expired"})
    if expected_kind is not None and data.get("kind") != expected_kind:
        raise DownloadTokenError("", params={"reason": "kind_mismatch", "expected": expected_kind})
    # 旧版(未绑定项目/用户)token 一律视为无效
    if data.get("project_id") is None or data.get("user_id") is None:
        raise DownloadTokenError("", params={"reason": "bad_payload"})
    return {
        "object_id": int(data["object_id"]),
        "kind": data.get("kind"),
        "project_id": int(data["project_id"]),
        "user_id": int(data["user_id"]),
    }


# ---------------------------------------------------------------------------
# 项目包导出(架构宪法 §12、domain-model §对象生命周期)
# ---------------------------------------------------------------------------


class PackageExport:
    """项目包导出结果(对象记录 + 下载授权 + 清单)。"""

    __slots__ = (
        "object_id", "oid", "sha256", "size_bytes", "media_type",
        "file_name", "manifest", "token", "expires_at",
    )

    def __init__(
        self,
        *,
        object_id: int,
        oid: str,
        sha256: str,
        size_bytes: int,
        media_type: str,
        file_name: str,
        manifest: dict,
        token: str,
        expires_at: datetime,
    ) -> None:
        self.object_id = object_id
        self.oid = oid
        self.sha256 = sha256
        self.size_bytes = size_bytes
        self.media_type = media_type
        self.file_name = file_name
        self.manifest = manifest
        self.token = token
        self.expires_at = expires_at

    def to_dict(self) -> dict[str, Any]:
        """API 序列化(不含包内容)。"""
        return {
            "object_id": self.object_id,
            "oid": self.oid,
            "sha256": self.sha256,
            "size_bytes": self.size_bytes,
            "media_type": self.media_type,
            "file_name": self.file_name,
            "token": self.token,
            "expires_at": self.expires_at.isoformat() if self.expires_at else None,
            "manifest": self.manifest,
        }


def _bound_dataset_ids(content: dict) -> list[int]:
    """从项目内容取出绑定的数据集版本 id 清单。"""
    return [
        int(binding["dataset_version_id"])
        for binding in content.get("dataset_bindings", [])
        if binding.get("dataset_version_id") is not None
    ]


def _collect_datasets(db: Session, dataset_version_ids: list[int]) -> list[dict]:
    """收集数据集版本(元数据 + 文件对象内容), 供打包使用。

    返回 [{"dataset": Dataset, "version": DatasetVersion,
            "files": [{"file": DatasetFile, "obj": dict(元数据), "content": bytes}]}]。
    """
    out: list[dict] = []
    for dvid in dict.fromkeys(int(i) for i in dataset_version_ids):
        version = db.get(DatasetVersion, dvid)
        if version is None:
            continue
        dataset = db.get(Dataset, version.dataset_id)
        files: list[dict] = []
        for f in db.execute(
            select(DatasetFile).where(DatasetFile.dataset_version_id == version.id)
        ).scalars():
            try:
                obj = object_info(db, f.object_id)
            except NotFoundError:
                continue
            files.append(
                {"file": f, "obj": obj, "content": get_object(db, obj["id"])}
            )
        out.append({"dataset": dataset, "version": version, "files": files})
    return out


def _collect_evidence(db: Session, project_id: int) -> list[dict]:
    """收集项目历史结果证据与评估引用(经任务归属项目，domain-model §快照、任务和结果)。"""
    packages = db.execute(
        select(EvidencePackage)
        .join(Task, Task.id == EvidencePackage.task_id)
        .where(Task.project_id == project_id)
        .order_by(EvidencePackage.id)
    ).scalars().all()
    out: list[dict] = []
    for pkg in packages:
        task = db.get(Task, pkg.task_id)
        snapshot = db.get(CalcSnapshot, pkg.calc_snapshot_id) if pkg.calc_snapshot_id else None
        assessments = db.execute(
            select(ResultAssessment)
            .where(ResultAssessment.evidence_package_id == pkg.id)
            .order_by(ResultAssessment.id)
        ).scalars().all()
        index = db.execute(
            select(ResultIndex).where(ResultIndex.evidence_package_id == pkg.id)
        ).scalars().all()
        content: bytes | None = None
        obj: dict | None = None
        try:
            obj = object_info(db, pkg.object_id)
            content = get_object(db, obj["id"])
        except NotFoundError:
            pass
        out.append(
            {
                "package": pkg, "task": task, "snapshot": snapshot,
                "assessments": assessments, "index": index,
                "object": obj, "content": content,
            }
        )
    return out


def _build_package_zip(
    db: Session, project: Project, draft: Draft, draft_content: dict,
) -> tuple[bytes, dict]:
    """组装项目包 zip 字节与清单(流式写入; 不含账号/权限/会话/密钥，见 domain-model §对象生命周期)。"""
    versions = db.execute(
        select(ProjectVersion)
        .where(ProjectVersion.project_id == project.id)
        .order_by(ProjectVersion.version_no)
    ).scalars().all()

    # 数据集版本 id 全集: 当前草稿绑定 + 全部版本绑定 + 证据快照绑定
    dataset_ids: list[int] = list(_bound_dataset_ids(draft_content))
    version_contents: dict[int, dict] = {}
    for version in versions:
        content = project_service.load_content_object(db, version.content_hash)
        version_contents[version.version_no] = content
        dataset_ids.extend(_bound_dataset_ids(content))
    evidence_list = _collect_evidence(db, project.id)
    for item in evidence_list:
        snapshot = item["snapshot"]
        if snapshot is not None:
            dataset_ids.extend(int(i) for i in (snapshot.dataset_version_ids or []))
    dataset_list = _collect_datasets(db, dataset_ids)

    objects_manifest: list[dict[str, Any]] = []
    files_meta: dict[str, Any] = {
        "draft": {"revision": draft.revision, "content_hash": draft.content_hash},
        "versions": [],
        "datasets": [],
        "evidence": [],
    }

    def _add(path: str, data: bytes, media_type: str) -> None:
        """写入 zip 条目并登记对象清单(路径/校验值/大小/类型)。"""
        objects_manifest.append(
            {"path": path, "sha256": sha256_hex(data), "size_bytes": len(data),
             "media_type": media_type}
        )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # 项目元数据(不导出所有者/创建者等账号信息)
        project_meta = {
            "id": project.id,
            "name": project.name,
            "description": project.description,
            "status": project.status,
            "currency": project.currency,
            "fixed_utc_offset_minutes": project.fixed_utc_offset_minutes,
            "schema_version": project.schema_version,
            "created_at": project.created_at,
        }
        project_json = json.dumps(jsonable(project_meta), ensure_ascii=False, indent=2).encode()
        _add("project.json", project_json, "application/json")
        zf.writestr("project.json", project_json)

        # 当前草稿(领域内容, 命令簿记不外泄)
        content = {k: v for k, v in draft_content.items() if k != "applied_commands"}
        draft_json = json.dumps(
            {"revision": draft.revision, "content_hash": draft.content_hash, "content": content},
            ensure_ascii=False, indent=2,
        ).encode()
        _add("draft.json", draft_json, "application/json")
        zf.writestr("draft.json", draft_json)

        # 项目版本(不可变, 版本内容 + 版本元数据; 不含创建者账号)
        for version in versions:
            path = f"versions/{version.version_no:04d}.json"
            version_doc = {
                "version": {
                    "version_no": version.version_no,
                    "name": version.name,
                    "description": version.description,
                    "reason": version.reason,
                    "fixed_utc_offset_minutes": version.fixed_utc_offset_minutes,
                    "currency": version.currency,
                    "schema_version": version.schema_version,
                    "content_hash": version.content_hash,
                    "created_at": version.created_at,
                    "source_draft_revision": version.source_draft_revision,
                },
                "content": version_contents.get(version.version_no, {}),
            }
            raw = json.dumps(jsonable(version_doc), ensure_ascii=False, indent=2).encode()
            _add(path, raw, "application/json")
            zf.writestr(path, raw)
            files_meta["versions"].append(path)

        # 数据集版本与溯源(数据本体 + 元数据 + 溯源/许可证/质量报告)
        for item in dataset_list:
            dataset = item["dataset"]
            version = item["version"]
            if dataset is None:
                continue
            base = f"datasets/{dataset.id}"
            meta_doc = {
                "dataset_id": dataset.id,
                "dataset_version_id": version.id,
                "dataset": {
                    "name": dataset.name,
                    "description": dataset.description,
                    "status": dataset.status,
                    "default_license": dataset.default_license,
                },
                "version": {
                    "version_no": version.version_no,
                    "timeline": version.timeline,
                    "resolution": version.resolution,
                    "fixed_utc_offset_minutes": version.fixed_utc_offset_minutes,
                    "fields": version.fields,
                    "units": version.units,
                    "quality_report": version.quality_report,
                    "provenance": version.provenance,
                    "license": version.license,
                    "content_hash": version.content_hash,
                    "created_at": version.created_at,
                    "created_reason": version.created_reason,
                },
                "files": [
                    {
                        "file_kind": item["file"].file_kind,
                        "format": item["file"].format,
                        "row_count": item["file"].row_count,
                        "size_bytes": item["file"].size_bytes,
                        "sha256": item["obj"]["sha256"],
                        "media_type": item["obj"]["media_type"],
                    }
                    for item in item["files"]
                ],
            }
            meta_raw = json.dumps(jsonable(meta_doc), ensure_ascii=False, indent=2).encode()
            meta_path = f"{base}/dataset.json"
            _add(meta_path, meta_raw, "application/json")
            zf.writestr(meta_path, meta_raw)
            files_meta["datasets"].append(meta_path)
            for file_item in item["files"]:
                f, obj, content_bytes = file_item["file"], file_item["obj"], file_item["content"]
                ext = "csv" if _FORMAT_BY_MEDIA.get(obj.get("media_type") or "", "") == "csv" else "json"
                entry = f"{base}/v{version.version_no}-{f.file_kind}.{ext}"
                _add(entry, content_bytes, obj.get("media_type") or "application/octet-stream")
                zf.writestr(entry, content_bytes)

        # 历史结果证据与评估引用(证据数据 + 评估四维结论; 不重新求解)
        for item in evidence_list:
            pkg = item["package"]
            task = item["task"]
            snapshot = item["snapshot"]
            base = f"evidence/{pkg.id}"
            evidence_doc = {
                "package": {
                    "id": pkg.id,
                    "status": pkg.status,
                    "content_hash": pkg.content_hash,
                    "created_at": pkg.created_at,
                },
                "task": {
                    "id": task.id if task else None,
                    "type": task.type if task else None,
                    "status": task.status if task else None,
                    "business_outcome": task.business_outcome if task else None,
                },
                "snapshot": {
                    "id": snapshot.id if snapshot else None,
                    "content_hash": snapshot.content_hash if snapshot else None,
                    "program_version": snapshot.program_version if snapshot else None,
                    "random_seed": snapshot.random_seed if snapshot else None,
                    "dataset_version_ids": (snapshot.dataset_version_ids if snapshot else []),
                    "canonical_assembly_text": (
                        snapshot.canonical_assembly_text if snapshot else None
                    ),
                    "assembly_sha256": snapshot.assembly_sha256 if snapshot else None,
                    "assembly_receipt": snapshot.assembly_receipt if snapshot else None,
                },
                "assessments": [
                    {
                        "id": a.id,
                        "assessor": a.assessor,
                        "dimension_physical": a.dimension_physical,
                        "dimension_optimality": a.dimension_optimality,
                        "dimension_financial": a.dimension_financial,
                        "dimension_reliability": a.dimension_reliability,
                        "overall_score": float(a.overall_score) if a.overall_score is not None else None,
                        "comment": a.comment,
                        "detail": a.detail,
                        "created_at": a.created_at,
                    }
                    for a in item["assessments"]
                ],
                "result_index": [
                    {
                        "id": r.id,
                        "result_hash": r.result_hash,
                        "assessment_id": r.assessment_id,
                        "is_latest": r.is_latest,
                        "created_at": r.created_at,
                    }
                    for r in item["index"]
                ],
            }
            evidence_raw = json.dumps(jsonable(evidence_doc), ensure_ascii=False, indent=2).encode()
            evidence_path = f"{base}.json"
            _add(evidence_path, evidence_raw, "application/json")
            zf.writestr(evidence_path, evidence_raw)
            files_meta["evidence"].append(evidence_path)
            if item["content"] is not None:
                media = (
                    (item["object"].get("media_type") if item["object"] else None)
                    or "application/octet-stream"
                )
                ext = "json" if "json" in media else "bin"
                entry = f"{base}/result.{ext}"
                _add(entry, item["content"], media)
                zf.writestr(entry, item["content"])

        # 清单最后写入(含全部条目校验值; 版本化清单: 格式版本/文件清单/对象清单)
        objects_manifest.sort(key=lambda e: e["path"])
        aggregate = sha256_hex(
            "".join(f"{e['path']}\0{e['sha256']}\0" for e in objects_manifest).encode("utf-8")
        )
        manifest: dict[str, Any] = {
            "format_version": PACKAGE_FORMAT_VERSION,
            "package_type": "project",
            "generated_at": datetime.now(UTC).isoformat(),
            "exporter": "iesplan",
            "exporter_version": __version__,
            "project": {k: v for k, v in project_meta.items() if k != "id"},
            "files": files_meta,
            "objects": objects_manifest,
            "checksums": {"entry_count": len(objects_manifest), "aggregate_sha256": aggregate},
        }
        manifest_raw = json.dumps(jsonable(manifest), ensure_ascii=False, indent=2).encode()
        zf.writestr("manifest.json", manifest_raw)
    return buf.getvalue(), manifest


def export_package(db: Session, user: User, project_id: int) -> PackageExport:
    """导出完整项目包(仅所有者，架构宪法 §12/domain-model §对象生命周期)。

    流程: 权限校验 → 组装 zip(模型/配置/草稿/版本/数据集版本与溯源/历史结果
    证据与评估引用/内容校验) → 内容寻址对象登记 →
    业务引用 + 审计 → 短期单对象下载授权。

    包内不含: 账号/权限与查看者名单/会话/全局系统配置/部署环境密钥。
    """
    project_service.ensure_access(db, user, project_id, "export_package")
    project = project_service.require_project(db, project_id)
    draft = project_service.get_current_draft(db, project)
    draft_content = project_service.load_content_object(db, draft.content_hash)
    zip_bytes, manifest = _build_package_zip(db, project, draft, draft_content)

    obj = put_object(
        db, zip_bytes, PACKAGE_MEDIA_TYPE, source_category="project_package",
    )
    add_ref(
        db, obj.id, "export_package", project.id,
        ref_entity_type="projects", purpose="项目包导出对象(23.2 保留)",
    )
    # ObjectHandle → 元数据 dict(公开门面统一形状)
    obj_info = object_info(db, obj.id)
    # 不再写 data_dir/packages 非托管副本(无引用/配额/校验/清理协议，架构宪法 §10);
    # 对象存储是包的唯一事实源, 下载经短期授权 token 走公开读取门面。

    audit_service.audit(
        db, user.id, audit_service.AUDIT_PROJECT_EXPORTED, "project", project.id,
        revision=draft.revision,
        result={"kind": "package", "package_object_id": obj.id, "size_bytes": len(zip_bytes)},
        checksum_info={"sha256": obj_info["sha256"]},
        extra={"file_name": f"project-package-{project.id}.zip"},
    )
    expires_at = datetime.now(UTC) + timedelta(seconds=DOWNLOAD_TOKEN_TTL_SECONDS)
    token = create_download_token(obj.id, "package", project_id=project.id, user_id=user.id)
    return PackageExport(
        object_id=obj.id, oid=obj_info["oid"], sha256=obj_info["sha256"],
        size_bytes=obj_info["size_bytes"], media_type=obj_info["media_type"],
        file_name=f"project-package-{project.id}.zip",
        manifest=manifest, token=token, expires_at=expires_at,
    )


# ---------------------------------------------------------------------------
# 项目包导入(校验 → 提案 → 确认; 每次导入创建新项目身份，domain-model §对象生命周期/§快照、任务和结果)
# ---------------------------------------------------------------------------


def _parse_package(data: bytes) -> tuple[dict, dict[str, bytes]]:
    """解析项目包 zip: (manifest, {entry_path: bytes}); 校验失败抛 ImportValidationError。

    前置门禁(任何解压读取之前执行, 防 ZIP Bomb 内存/CPU 耗尽):
    - 上传字节上限(MAX_PACKAGE_BYTES, 2GB);
    - zip 文件头与条目数预检(MAX_PACKAGE_ENTRIES, 5000);
    - 单条目解压大小(按 zip 头声明的 file_size 预检, MAX_PACKAGE_ENTRY_BYTES);
    - 总解压大小上限(MAX_PACKAGE_TOTAL_BYTES)。
    超限抛 PackageSizeError(PKG-SIZE-001, 413)。

    校验项:
    - 格式: 合法 zip, 无路径穿越条目;
    - 兼容性: 主版本号与当前格式兼容(1.x);
    - 清单: manifest.json 存在, package_type=project, 无账号/权限/会话/全局配置/密钥;
    - 完整性: 对象清单逐对象 sha256 + 大小校验, 包内文件与清单一一对应;
    - 必需文件: project.json / draft.json 存在。
    """
    reasons: list[str] = []
    if len(data) > MAX_PACKAGE_BYTES:
        raise PackageSizeError(
            "",
            params={"reason": "package_too_large", "max_bytes": MAX_PACKAGE_BYTES,
                    "actual_bytes": len(data)},
        )
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except (zipfile.BadZipFile, OSError) as exc:
        raise ImportValidationError(["文件不是合法的 zip 项目包"]) from exc
    with zf:
        # 条目数与单条目/总解压大小预检(zip 头部声明值, 不做真实解压)
        infos = zf.infolist()
        if len(infos) > MAX_PACKAGE_ENTRIES:
            raise PackageSizeError(
                "",
                params={"reason": "too_many_entries", "max_entries": MAX_PACKAGE_ENTRIES,
                        "actual_entries": len(infos)},
            )
        total_uncompressed = 0
        for info in infos:
            if info.is_dir():
                continue
            if info.file_size > MAX_PACKAGE_ENTRY_BYTES:
                raise PackageSizeError(
                    "",
                    params={"reason": "entry_too_large", "entry": info.filename,
                            "max_bytes": MAX_PACKAGE_ENTRY_BYTES, "actual_bytes": info.file_size},
                )
            total_uncompressed += info.file_size
            if total_uncompressed > MAX_PACKAGE_TOTAL_BYTES:
                raise PackageSizeError(
                    "",
                    params={"reason": "total_uncompressed_too_large",
                            "max_bytes": MAX_PACKAGE_TOTAL_BYTES, "actual_bytes": total_uncompressed},
                )
        entries: dict[str, bytes] = {}
        for info in infos:
            if info.is_dir():
                continue
            name = info.filename
            if name.startswith("/") or ".." in name.split("/"):
                reasons.append(f"条目路径非法(路径穿越): {name}")
                continue
            entries[name] = zf.read(info)
    if reasons:
        raise ImportValidationError(reasons)
    if "manifest.json" not in entries:
        raise ImportValidationError(["包内缺少 manifest.json 清单"])
    try:
        manifest = json.loads(entries["manifest.json"].decode("utf-8"))
    except (UnicodeDecodeError, ValueError) as exc:
        raise ImportValidationError(["manifest.json 无法解析"]) from exc
    if not isinstance(manifest, dict):
        raise ImportValidationError(["manifest.json 结构非法"])

    # 兼容性: 格式主版本
    format_version = manifest.get("format_version")
    if not isinstance(format_version, str) or not format_version.split(".", 1)[0] == "1":
        reasons.append(f"格式版本不兼容: {format_version!r}, 期望 1.x")
    if manifest.get("package_type") != "project":
        reasons.append(f"包类型非法: {manifest.get('package_type')!r}, 期望 project")
    # 禁止携带账号/权限/会话/全局配置/密钥
    forbidden = FORBIDDEN_MANIFEST_KEYS & set(manifest)
    if forbidden:
        reasons.append(f"包包含禁止内容({', '.join(sorted(forbidden))}), 拒绝导入")

    # 完整性: 对象清单逐对象校验(sha256 + 大小)
    objects_manifest = manifest.get("objects")
    if not isinstance(objects_manifest, list):
        reasons.append("清单缺少 objects 对象清单")
    else:
        seen_paths: set[str] = set()
        for entry in objects_manifest:
            if not isinstance(entry, dict):
                reasons.append("对象清单条目非法")
                continue
            path = entry.get("path")
            expected_sha = entry.get("sha256")
            expected_size = entry.get("size_bytes")
            if not isinstance(path, str) or path not in entries:
                reasons.append(f"对象清单条目缺失文件: {path}")
                continue
            if path in seen_paths:
                reasons.append(f"对象清单重复条目: {path}")
                continue
            seen_paths.add(path)
            actual = entries[path]
            if not isinstance(expected_sha, str) or sha256_hex(actual) != expected_sha:
                reasons.append(f"对象校验值不符(sha256): {path}")
            if not isinstance(expected_size, int) or len(actual) != expected_size:
                reasons.append(f"对象大小不符: {path}")
        # 反向: 包内文件(除清单)必须全部在对象清单中
        for path in entries:
            if path == "manifest.json" or path in seen_paths:
                continue
            reasons.append(f"包内文件未在对象清单中: {path}")

    for required in REQUIRED_PACKAGE_FILES:
        if required not in entries:
            reasons.append(f"缺少必需文件: {required}")
    if reasons:
        raise ImportValidationError(reasons)
    return manifest, entries


def _unique_project_name(db: Session, base: str) -> str:
    """项目名称去重: 已存在同名项目时追加 " (导入 n)" 后缀(不静默覆盖)。"""
    candidate = base
    index = 2
    while db.execute(select(Project.id).where(Project.name == candidate)).first() is not None:
        candidate = f"{base} (导入 {index})"
        index += 1
    return candidate


def import_proposal(
    db: Session,
    user: User,
    file_bytes: bytes,
    idempotency_key: str | None = None,
) -> ImportProposal:
    """创建导入提案: 校验 → 暂存对象 → 拟创建项目快照 → 校验报告(domain-model §对象生命周期)。

    - 校验失败(格式/兼容性/清单/完整性)抛 ImportValidationError, 不创建任何记录;
    - 相同源文件(sha256)同一提议人已有提案时幂等返回既有提案(不重复暂存);
    - 暂存: 包内全部对象写入内容寻址对象存储(按 sha256 去重);
    - 拟创建项目快照: 同事务创建新项目身份(导入者即所有者, 原授权不迁移),
      校验报告与分区提交内容写入 review_summary(review_errors 空);
    - 确认导入见 confirm_import。
    """
    manifest, entries = _parse_package(file_bytes)
    source_hash = sha256_hex(file_bytes)

    # 幂等: 同一提议人 + 同一源文件 → 返回既有提案(内容寻址, 不重复暂存)
    existing = db.execute(
        select(ImportProposal)
        .where(
            ImportProposal.proposer_id == user.id,
            ImportProposal.source_hash == source_hash,
        )
        .order_by(ImportProposal.id.desc())
        .limit(1)
    ).scalar_one_or_none()
    if existing is not None:
        return existing

    # 1) 暂存对象: 包内全部对象内容寻址落盘(校验已通过, 逐对象 sha256 一致)
    staged: dict[str, dict] = {}  # path → 元数据 dict(公开门面)
    for entry in manifest.get("objects", []):
        path = entry["path"]
        staged[path] = put_object(
            db, entries[path], entry.get("media_type") or "application/octet-stream",
            source_category="project_package_import",
        )
    source_obj = put_object(
        db, file_bytes, PACKAGE_MEDIA_TYPE, source_category="import_package_source",
    )

    # 2) 拟创建项目快照: 新项目身份(每次导入新身份, 导入者成为所有者)
    project_meta = manifest.get("project") or {}
    name = _unique_project_name(db, str(project_meta.get("name") or "导入项目"))
    currency = project_meta.get("currency") or "CNY"
    if currency not in ("CNY", "USD"):
        raise ImportValidationError([f"包内币种非法: {currency}"])
    try:
        offset = int(project_meta.get("fixed_utc_offset_minutes", 480))
    except (TypeError, ValueError) as exc:
        raise ImportValidationError(["包内 UTC 偏移非法"]) from exc
    if not -720 <= offset <= 840:
        raise ImportValidationError([f"包内 UTC 偏移越界: {offset}"])
    project = Project(
        name=name,
        description=project_meta.get("description"),
        status="active",
        owner_id=user.id,
        currency=currency,
        fixed_utc_offset_minutes=offset,
        schema_version=int(project_meta.get("schema_version", 1) or 1),
        created_by=user.id,
    )
    db.add(project)
    db.flush()

    # 3) 导入提案(校验报告 + 分区提交内容, 01 §10.4)
    # 0.4.0: 不再写 source_path(该列可空且无任何消费点) —— storage_path 属
    # §11 内部路径, 不得进入审计记录; 可追溯性由 review_summary.source_object_id
    # (内容寻址对象 ID)与 source_hash 承担。
    proposal = ImportProposal(
        project_id=project.id,
        proposer_id=user.id,
        source_type="json",
        source_hash=source_hash,
        status="proposed",
        review_summary={
            "package": {
                "format_version": manifest.get("format_version"),
                "generated_at": manifest.get("generated_at"),
                "exporter": manifest.get("exporter"),
                "exporter_version": manifest.get("exporter_version"),
            },
            "project_snapshot": {
                "name": name, "currency": currency,
                "fixed_utc_offset_minutes": offset,
                "schema_version": project.schema_version,
            },
            "staging": {
                "object_count": len(staged),
                "objects": [
                    {"path": path, "object_id": obj.id, "sha256": obj.sha256}
                    for path, obj in sorted(staged.items())
                ],
                "source_object_id": source_obj.id,
            },
            "checks": {
                "zip_ok": True, "manifest_ok": True, "integrity_ok": True,
                "integrity_verified_objects": len(manifest.get("objects", [])),
            },
            "idempotency_key": idempotency_key,
        },
        review_errors={},
    )
    db.add(proposal)
    db.flush()
    audit_service.audit(
        db, user.id, audit_service.AUDIT_PROJECT_IMPORT_PROPOSED, "import_proposals",
        proposal.id,
        result={"project_id": project.id, "source_hash": source_hash,
                "staged_objects": len(staged)},
        checksum_info={"sha256": source_hash},
    )
    return proposal


def _create_draft_row(db: Session, project: Project, content: dict, user: User) -> Draft:
    """新项目身份创建初始草稿(revision=1, 与版本服务同构的内容寻址路径)。"""
    content_hash = project_service.store_content_object(db, content)
    draft = Draft(
        project_id=project.id, revision=1, content_hash=content_hash,
        parent_draft_id=None, is_current=True, updated_by=user.id,
    )
    db.add(draft)
    db.flush()
    project.current_draft_id = draft.id
    return draft


def _object_by_sha256(db: Session, digest: str) -> dict:
    """按 sha256 取对象(暂存对象查找, STO-05: 经公开门面返回元数据 dict)。"""
    try:
        return object_by_sha256(db, digest)
    except NotFoundError as exc:
        raise AppError(
            "导入暂存对象缺失(数据损坏)",
            code="PKG-IMP-002", severity=SEVERITY_ERROR,
            message_key="ies.diag.store.corrupt",
            params={"sha256": digest},
        ) from exc


def confirm_import(db: Session, user: User, proposal_id: int) -> Project:
    """提交导入(U14, domain-model §对象生命周期): 创建新项目身份, 导入者成为所有者。

    分区提交内容(单事务):
    - 数据集: 重建 Dataset/DatasetVersion/DatasetFile(引用暂存对象, 原标识重映射);
    - 草稿: revision=1 领域内容(数据集绑定重映射, 证据来源登记);
    - 版本: 按包内版本顺序重建 ProjectVersion(新身份版本号, 不倒写原版本);
    - 证据: 历史结果作为证据来源保留 — 登记对象引用(imported_evidence)与
      评估摘要, 不创建本地任务(不伪造本地任务, domain-model §快照、任务和结果)。

    导入约束: 不得静默覆盖(名称去重 + 新项目身份); 账号/权限/会话不随包导入;
    导入者成为新项目所有者; 原授权关系不迁移。
    """
    proposal = db.get(ImportProposal, proposal_id)
    if proposal is None:
        raise NotFoundError(
            "导入提案不存在",
            params={"proposal_id": proposal_id},
            location={"object_type": "import_proposals", "object_id": proposal_id},
        )
    if proposal.proposer_id != user.id:
        raise ForbiddenError(
            "仅提案人可确认导入", params={"proposal_id": proposal_id},
            location={"object_type": "import_proposals", "object_id": proposal_id},
        )
    project = db.get(Project, proposal.project_id)
    if project is None:
        raise NotFoundError("导入提案关联项目缺失", params={"proposal_id": proposal_id})
    if proposal.status == "applied":
        return project  # 幂等重放: 已导入则返回导入结果项目
    if proposal.status == "rejected":
        raise ConflictError(
            "导入提案已被拒绝", params={"proposal_id": proposal_id, "status": proposal.status}
        )
    if proposal.status not in ("proposed", "validated", "approved"):
        raise ConflictError(
            "导入提案状态不允许确认", params={"proposal_id": proposal_id, "status": proposal.status}
        )

    # 复核源包(完整性), 取分区提交内容
    summary = proposal.review_summary or {}
    source_object_id = (summary.get("staging") or {}).get("source_object_id")
    file_bytes = get_object(db, int(source_object_id))
    manifest, entries = _parse_package(file_bytes)

    # 1) 数据集(先建, 供绑定重映射; 原数据集版本标识 → 新标识)
    dataset_id_map: dict[int, int] = {}
    for meta_path in manifest.get("files", {}).get("datasets", []):
        if meta_path not in entries:
            continue
        meta = json.loads(entries[meta_path].decode("utf-8"))
        ds_meta = meta.get("dataset") or {}
        ver_meta = meta.get("version") or {}
        dataset = Dataset(
            project_id=project.id,
            name=str(ds_meta.get("name") or "导入数据集"),
            description=ds_meta.get("description"),
            status="published",
            default_license=ds_meta.get("default_license"),
            created_by=user.id,
        )
        db.add(dataset)
        db.flush()
        version = DatasetVersion(
            dataset_id=dataset.id,
            version_no=int(ver_meta.get("version_no", 1)),
            timeline=str(ver_meta.get("timeline") or "hourly"),
            resolution=ver_meta.get("resolution"),
            fixed_utc_offset_minutes=int(ver_meta.get("fixed_utc_offset_minutes", 480)),
            fields=ver_meta.get("fields") or {},
            units=ver_meta.get("units") or {},
            quality_report=ver_meta.get("quality_report"),
            provenance=ver_meta.get("provenance"),
            license=ver_meta.get("license"),
            content_hash=str(ver_meta.get("content_hash") or ""),
            created_by=user.id,
            created_reason="imported",
        )
        db.add(version)
        db.flush()
        dataset_id_map[int(meta.get("dataset_version_id") or 0)] = version.id
        for file_meta in meta.get("files", []):
            obj = _object_by_sha256(db, str(file_meta["sha256"]))
            db.add(
                DatasetFile(
                    dataset_version_id=version.id,
                    object_id=obj["id"],
                    file_kind=str(file_meta.get("file_kind") or "data"),
                    format=str(file_meta.get("format") or "csv"),
                    row_count=int(file_meta.get("row_count", 0)),
                    size_bytes=int(file_meta.get("size_bytes", 0)),
                )
            )

    def _remap(content: dict) -> dict:
        """重映射数据集绑定到新项目的数据集版本标识(原授权/标识不迁移)。"""
        bindings = content.get("dataset_bindings") or []
        for binding in bindings:
            original = binding.get("dataset_version_id")
            if original in dataset_id_map:
                binding["dataset_version_id"] = dataset_id_map[original]
                binding.pop("dataset_id", None)
        content.pop("applied_commands", None)  # 命令簿记不外泄
        return content

    # 2) 证据来源: 历史结果作为证据来源保留(不伪造本地任务, domain-model §快照、任务和结果)
    imported_evidence: list[dict[str, Any]] = []
    for evidence_path in manifest.get("files", {}).get("evidence", []):
        if evidence_path not in entries:
            continue
        doc = json.loads(entries[evidence_path].decode("utf-8"))
        pkg_meta = doc.get("package") or {}
        task_meta = doc.get("task") or {}
        snapshot_meta = doc.get("snapshot") or {}
        assessments = doc.get("assessments") or []
        base = evidence_path.removesuffix(".json")
        ref_objects = [
            entry for entry in manifest.get("objects", [])
            if entry.get("path") == evidence_path or entry.get("path", "").startswith(f"{base}/")
        ]
        for entry in ref_objects:
            obj = _object_by_sha256(db, entry["sha256"])
            add_ref(
                db, obj["id"], "imported_evidence", project.id,
                ref_entity_type="projects",
                purpose="导入的历史结果证据来源(不伪造本地任务)",
            )
        imported_evidence.append(
            {
                "package_id": pkg_meta.get("id"),
                "content_hash": pkg_meta.get("content_hash"),
                "task": {k: task_meta.get(k) for k in ("type", "status", "business_outcome")},
                "snapshot_content_hash": snapshot_meta.get("content_hash"),
                "assessments": [
                    {k: a.get(k) for k in ("id", "assessor", "dimension_physical",
                                           "dimension_optimality", "dimension_financial",
                                           "dimension_reliability", "overall_score", "comment")}
                    for a in assessments
                ],
                "objects": [
                    {"path": e["path"], "sha256": e["sha256"]} for e in ref_objects
                ],
            }
        )

    # 3) 草稿(revision=1): 领域内容 + 证据来源登记
    draft_doc = json.loads(entries["draft.json"].decode("utf-8"))
    draft_content = _remap(dict(draft_doc.get("content") or {}))
    draft_content["imported_evidence"] = imported_evidence
    _create_draft_row(db, project, draft_content, user)

    # 4) 版本: 按包内版本顺序重建(新身份版本号, parent 链按导入顺序)
    prev_version: ProjectVersion | None = None
    version_paths = sorted(manifest.get("files", {}).get("versions", []))
    for version_path in version_paths:
        if version_path not in entries:
            continue
        doc = json.loads(entries[version_path].decode("utf-8"))
        ver_meta = doc.get("version") or {}
        version_content = _remap(dict(doc.get("content") or {}))
        content_hash = project_service.store_content_object(db, version_content)
        db.flush()  # 内容重映射可能产生新对象行, 先 flush 再按校验值取对象
        obj = _object_by_sha256(db, content_hash)
        version = ProjectVersion(
            project_id=project.id,
            version_no=int(ver_meta.get("version_no", 1)),
            name=str(ver_meta.get("name") or f"导入版本 {ver_meta.get('version_no', 1)}"),
            description=ver_meta.get("description"),
            created_by=user.id,
            parent_version_id=prev_version.id if prev_version is not None else None,
            source_draft_id=None,
            source_draft_revision=None,
            reason="imported",
            fixed_utc_offset_minutes=int(
                ver_meta.get("fixed_utc_offset_minutes", project.fixed_utc_offset_minutes)
            ),
            currency=ver_meta.get("currency") or project.currency,
            schema_version=int(ver_meta.get("schema_version", 1) or 1),
            content_hash=content_hash,
        )
        db.add(version)
        db.flush()
        db.add(
            VersionRef(
                project_version_id=version.id,
                ref_type="object",
                object_id=obj["id"],
                ref_key="project_version_content",
                ref_hash=content_hash,
            )
        )
        prev_version = version
    if prev_version is not None:
        project.current_version_id = prev_version.id

    # 5) 提案收尾 + 审计
    proposal.status = "applied"
    proposal.decided_by = user.id
    proposal.decided_at = datetime.now(UTC)
    audit_service.audit(
        db, user.id, audit_service.AUDIT_PROJECT_IMPORTED, "project", project.id,
        revision=1,
        result={
            "source_hash": proposal.source_hash,
            "imported_versions": len(version_paths),
            "imported_datasets": len(dataset_id_map),
            "evidence_objects": len(imported_evidence),
        },
        checksum_info={"sha256": proposal.source_hash},
    )
    return project


# ---------------------------------------------------------------------------
# Excel 报告导出(U15/U14, domain-model §快照、任务和结果 / contracts §公共文件契约 / REQ-EXPORT-001: 固定模板, 固定引用, 不重新求解)
# ---------------------------------------------------------------------------


def _parse_evidence_content(raw: bytes | None) -> dict | None:
    """解析证据对象内容(JSON 字典), 解析失败返回 None。"""
    if not raw:
        return None
    try:
        parsed = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, ValueError):
        return None
    return parsed if isinstance(parsed, dict) else None


def _first_section(content: dict | None, *keys: str) -> Any:
    """按键名顺序取证据内容中的摘要节(未命中返回 None)。"""
    if not content:
        return None
    for key in keys:
        if key in content:
            return content[key]
    return None


def _section_rows(section: Any) -> list[tuple[str, Any]]:
    """摘要节 → (名称, 值) 行(兼容 dict 与 [{name,value,unit}] 两种形态)。"""
    rows: list[tuple[str, Any]] = []
    if isinstance(section, dict):
        rows = [(str(k), v) for k, v in section.items()]
    elif isinstance(section, list):
        for item in section:
            if not isinstance(item, dict):
                continue
            name = item.get("name") or item.get("key") or item.get("label")
            if name is None:
                continue
            value = item.get("value")
            unit = item.get("unit")
            rows.append((str(name), f"{value} {unit}" if unit is not None else value))
    return rows


def _excel_safe(value: Any) -> Any:
    """Excel 公式注入防护(M-09): 用户可控字符串以 = + - @ 开头时前置单引号。

    所有写入 Excel 单元格的用户可控值都必须经过本函数, 防止恶意项目名/设备名/
    评论等在打开报表时被 Excel/LibreOffice 当作公式执行(客户端文件风险)。
    """
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _set_sheet_title(ws, title_zh: str, title_en: str) -> None:
    """写入双语节标题(加粗)。"""
    cell = ws.cell(row=1, column=1, value=_excel_safe(f"{title_zh} / {title_en}"))
    cell.font = Font(bold=True)


def _write_kv(ws, rows: list[tuple[str, Any]], start_row: int = 3) -> int:
    """写键值行(两列), 返回下一可用行号(全部值经 _excel_safe 防公式注入)。"""
    row = start_row
    for key, value in rows:
        ws.cell(row=row, column=1, value=_excel_safe(str(key)))
        if value is None:
            ws.cell(row=row, column=2, value="—")
        elif isinstance(value, (dict, list)):
            ws.cell(row=row, column=2, value=_excel_safe(json.dumps(value, ensure_ascii=False)[:500]))
        else:
            ws.cell(row=row, column=2, value=_excel_safe(value))
        row += 1
    return row


def export_excel(
    db: Session,
    user: User,
    project_id: int,
    evidence_package_id: int,
    assessment_id: int,
    lang: str = "zh",
) -> bytes:
    """导出固定模板 Excel 报告(查看者可导出, domain-model §快照、任务和结果 / contracts §公共文件契约 / REQ-EXPORT-001)。

    - 固定引用给定证据包与结果评估, 导出时不重新求解(11.2);
    - 标题中英双语(默认简体中文); 内容: 项目版本/计算快照/数据版本/计算配置/
      算法/结果状态/四维结论/主要指标表/设备配置/财务摘要/环境摘要/工程摘要/
      适用范围与限制;
    - 注明适用单位与数据来源(数据集版本/溯源/许可证/内容校验值)。
    """
    project_service.ensure_access(db, user, project_id, "export_excel")
    project = project_service.require_project(db, project_id)
    evidence = db.get(EvidencePackage, evidence_package_id)
    if evidence is None:
        raise NotFoundError(
            "证据包不存在", params={"evidence_package_id": evidence_package_id},
            location={"object_type": "evidence_packages", "object_id": evidence_package_id},
        )
    task = db.get(Task, evidence.task_id)
    if task is None or task.project_id != project_id:
        raise NotFoundError(
            "证据包不属于该项目", params={"evidence_package_id": evidence_package_id},
        )
    assessment = db.get(ResultAssessment, assessment_id)
    if assessment is None or assessment.evidence_package_id != evidence.id:
        raise NotFoundError(
            "结果评估不存在或与证据包不匹配",
            params={"assessment_id": assessment_id, "evidence_package_id": evidence.id},
            location={"object_type": "result_assessments", "object_id": assessment_id},
        )
    snapshot = db.get(CalcSnapshot, evidence.calc_snapshot_id)
    version = db.get(ProjectVersion, project.current_version_id) if project.current_version_id else None
    version_content: dict = {}
    if version is not None:
        version_content = project_service.load_content_object(db, version.content_hash)
    evidence_content = _parse_evidence_content(
        get_object(db, evidence.object_id)
    )

    # 数据版本(计算快照绑定的数据集版本 + 溯源/许可证)
    dataset_rows: list[dict[str, Any]] = []
    dvid_list = (snapshot.dataset_version_ids if snapshot else None) or []
    for dvid in dvid_list:
        dver = db.get(DatasetVersion, int(dvid))
        if dver is None:
            continue
        dset = db.get(Dataset, dver.dataset_id)
        dataset_rows.append(
            {
                "dataset": dset.name if dset else dver.dataset_id,
                "version_no": dver.version_no,
                "resolution": dver.resolution,
                "content_hash": dver.content_hash,
                "provenance": dver.provenance,
                "license": dver.license,
            }
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "报告总览"
    # 固定标题: 中英双语(默认简体中文在前)
    title = ws.cell(row=1, column=1, value="pIES 项目结果报告 / pIES Project Result Report")
    title.font = Font(bold=True, size=14)
    lang_label = "简体中文" if lang != "en" else "English"
    ws.cell(row=2, column=1, value=f"生成语言: {lang_label} / Language: {lang}")

    overview: list[tuple[str, Any]] = []
    overview.append(("项目名称 / Project name", project.name))
    overview.append(("项目状态 / Project status", project.status))
    overview.append(("币种 / Currency", project.currency))
    if version is not None:
        overview.append(("项目版本 / Project version", f"#{version.version_no} {version.name}"))
        overview.append(("版本说明 / Version reason", version.reason))
        overview.append(("版本内容校验 / Version content hash", version.content_hash))
    else:
        overview.append(("项目版本 / Project version", "—"))
    if snapshot is not None:
        overview.append(("计算快照 / Calc snapshot", f"#{snapshot.id}"))
        overview.append(("快照内容校验 / Snapshot content hash", snapshot.content_hash))
        overview.append(("程序版本 / Program version", snapshot.program_version))
        overview.append(("随机种子 / Random seed", snapshot.random_seed))
    else:
        overview.append(("计算快照 / Calc snapshot", "—"))
    overview.append(("数据版本 / Data versions", len(dataset_rows)))
    calc_config = snapshot.calc_config_snapshot if snapshot else {}
    overview.append(("计算配置 / Calc config", f"{len(calc_config.get('params') or {})} 参数 / parameters"))
    algorithm = calc_config.get("algorithm") or calc_config.get("solver")
    overview.append(("算法 / Algorithm", algorithm or "—"))
    overview.append(("结果状态 / Result status", task.status))
    overview.append(("业务结局 / Business outcome", task.business_outcome or "—"))
    overview.append(("证据包状态 / Evidence status", evidence.status))
    overview.append(("关联标识 / Reference", f"evidence_package={evidence.id}, assessment={assessment.id}"))
    overview.append(
        (
            "四维结论 / Four-dimension conclusion",
            f"物理 {assessment.dimension_physical} / 最优 {assessment.dimension_optimality} / "
            f"财务 {assessment.dimension_financial} / 可靠 {assessment.dimension_reliability}"
            + (
                f" | 综合评分 {float(assessment.overall_score)}"
                if assessment.overall_score is not None else ""
            ),
        )
    )
    overview.append(("评估意见 / Assessment comment", assessment.comment or "—"))
    _write_kv(ws, overview)
    row = 3 + len(overview) + 1
    # 适用范围与限制(固定引用证据与评估, 不重新求解)
    applicability = _first_section(evidence_content, "applicability", "scope", "适用")
    ws.cell(row=row, column=1, value="适用范围与限制 / Applicability and limitations").font = Font(bold=True)
    _write_kv(ws, _section_rows(applicability) or [("适用范围", "—")], start_row=row + 1)
    ws.cell(
        row=row + 6, column=1,
        value="本报告固定引用证据包与结果评估, 导出时不重新求解 / "
              "This report references a fixed evidence package and assessment; no re-solve on export.",
    ).font = Font(italic=True)
    ws.cell(
        row=row + 7, column=1,
        value="适用单位 / Units: 指标单位以各摘要表注记为准(如 kWh、MW、°C、元/kWh、kgCO₂/kWh)。",
    )
    ws.cell(
        row=row + 8, column=1,
        value="数据来源 / Data sources: " + ("; ".join(
            f"{d['dataset']} v{d['version_no']}({d['resolution']}, 校验 {str(d['content_hash'])[:12]}…, "
            f"许可证 {d['license'] or '—'})"
            for d in dataset_rows
        ) or "无绑定数据版本"),
    )

    # 主要指标表
    kpis = _first_section(evidence_content, "kpis", "key_metrics", "metrics")
    kp_sheet = wb.create_sheet("主要指标")
    _set_sheet_title(kp_sheet, "主要指标表", "Key Metrics Table")
    kp_rows = kpis if isinstance(kpis, list) else _section_rows(kpis)
    r = 3
    if kp_rows:
        kp_sheet.cell(row=2, column=1, value="指标 / Metric")
        kp_sheet.cell(row=2, column=2, value="值 / Value")
        kp_sheet.cell(row=2, column=3, value="单位 / Unit")
        for item in kp_rows:
            if isinstance(item, dict):
                kp_sheet.cell(
                    row=r, column=1,
                    value=_excel_safe(str(item.get("name") or item.get("key") or "")),
                )
                kp_sheet.cell(row=r, column=2, value=_excel_safe(item.get("value")))
                kp_sheet.cell(row=r, column=3, value=_excel_safe(item.get("unit") or ""))
            elif isinstance(item, tuple) and len(item) >= 2:
                kp_sheet.cell(row=r, column=1, value=_excel_safe(str(item[0])))
                kp_sheet.cell(row=r, column=2, value=_excel_safe(item[1]))
            r += 1
    else:
        kp_sheet.cell(row=3, column=1, value="无指标数据 / No metric data")

    # 设备配置(来自版本内容模型, 只展示不重算, REQ-RESULT-002)
    dev_sheet = wb.create_sheet("设备配置")
    _set_sheet_title(dev_sheet, "设备配置", "Equipment Configuration")
    devices = version_content.get("model", {}).get("devices", [])
    if devices:
        dev_sheet.cell(row=2, column=1, value="名称 / Name")
        dev_sheet.cell(row=2, column=2, value="类型 / Type")
        dev_sheet.cell(row=2, column=3, value="参数 / Parameters")
        r = 3
        for dev in devices:
            params = dev.get("params") or {}
            type_name = params.get("type_detail") or dev.get("device_type") or "—"
            dev_sheet.cell(row=r, column=1, value=_excel_safe(str(dev.get("name") or "")))
            dev_sheet.cell(row=r, column=2, value=_excel_safe(str(type_name)))
            dev_sheet.cell(
                row=r, column=3,
                value=_excel_safe(json.dumps(params, ensure_ascii=False)[:300]),
            )
            r += 1
    else:
        dev_sheet.cell(row=3, column=1, value="无设备配置 / No equipment")

    # 财务/环境/工程摘要(证据包内容, 固定引用, 不重新求解)
    for sheet_name, title_zh, title_en, section in (
        ("财务摘要", "财务摘要", "Financial Summary",
         _first_section(evidence_content, "financial", "finance")),
        ("环境摘要", "环境摘要", "Environmental Summary",
         _first_section(evidence_content, "environmental", "environment", "emissions")),
        ("工程摘要", "工程摘要", "Engineering Summary",
         _first_section(evidence_content, "engineering", "engineering_summary")),
    ):
        ws_s = wb.create_sheet(sheet_name)
        _set_sheet_title(ws_s, title_zh, title_en)
        rows = _section_rows(section)
        _write_kv(ws_s, rows if rows else [("无摘要数据 / No summary data", "—")])

    buf = io.BytesIO()
    wb.save(buf)
    audit_service.audit(
        db, user.id, audit_service.AUDIT_PROJECT_EXPORTED, "project", project.id,
        result={"kind": "excel", "evidence_package_id": evidence.id, "assessment_id": assessment.id},
        checksum_info={"content_hash": evidence.content_hash},
        extra={"lang": lang},
    )
    return buf.getvalue()


__all__ = [
    "ImportValidationError",
    "DownloadTokenError",
    "PackageExport",
    "PACKAGE_FORMAT_VERSION",
    "DOWNLOAD_TOKEN_TTL_SECONDS",
    "create_download_token",
    "verify_download_token",
    "export_package",
    "import_proposal",
    "confirm_import",
    "export_excel",
]
