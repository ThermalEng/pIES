"""项目包(U15)与导出 API 集成测试。

覆盖流程:
- 导出项目包(仅所有者; 含版本化清单/对象清单/内容校验)与下载;
- 导入: 校验通过 → 创建新项目身份(所有者=导入者, 名称不覆盖)→ 历史结果
  作为证据来源保留(不伪造本地任务)→ 幂等重放;
- 导入校验失败拒绝(篡改内容/非 zip);
- Excel 导出(查看者可导出; 标题中英双语; 固定引用证据包与评估)与下载;
- 查看者可导 Excel、不可导项目包; 非成员 403;
- 短期单对象授权过期/伪造拒绝;
- 管理端点: 审计查询/解锁任务/停用所有者转移/健康/存储(管理员 403 边界)。

测试环境: SQLite :memory:(StaticPool 共享连接) + tmp 对象存储目录 + 内存队列,
不依赖部署 Postgres/Redis; 通过 app.dependency_overrides 替换 get_db。
"""

from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import os
import time
import zipfile
from collections.abc import Iterator
from datetime import UTC, datetime, timedelta
from pathlib import Path
from uuid import uuid4

# 单文件运行时的安全网: 固定 SQLite + 内存队列, 避免误连部署 Postgres/Redis
os.environ.setdefault("IESPLAN_DB_URL", "sqlite+pysqlite://")
os.environ.setdefault("IESPLAN_QUEUE", "memory")

import pytest  # noqa: E402
from auth_helpers import login_headers, make_user  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import create_engine, select  # noqa: E402
from sqlalchemy.engine import Engine  # noqa: E402
from sqlalchemy.orm import Session, sessionmaker  # noqa: E402
from sqlalchemy.pool import StaticPool  # noqa: E402

from iesplan.api import admin as admin_api  # noqa: E402
from iesplan.api import exports as exports_api  # noqa: E402
from iesplan.api import projects as projects_api  # noqa: E402
from iesplan.config import settings  # noqa: E402
from iesplan.core.errors import ForbiddenError  # noqa: E402
from iesplan.core.idgen import sha256_hex  # noqa: E402
from iesplan.db import Base, get_db  # noqa: E402
from iesplan.main import create_app  # noqa: E402
from iesplan.models.audit import AuditLog, ImportProposal
from iesplan.models.calc import CalcSnapshot, ComputeSlot, Task, TaskAttempt, TaskLease  # noqa: E402
from iesplan.models.dataset import Dataset, DatasetFile, DatasetVersion  # noqa: E402
from iesplan.models.identity import User  # noqa: E402
from iesplan.models.project import (  # noqa: E402
    AdminMaintenanceAction,
    Project,
)
from iesplan.models.result import EvidencePackage, ResultAssessment, ResultIndex  # noqa: E402
from iesplan.services import package as package_service  # noqa: E402
from iesplan.services import project as project_service  # noqa: E402
from iesplan.services import queue  # noqa: E402
from iesplan.storage import put_object
from iesplan.storage.persistence import ObjectRef  # noqa: E402

# ---------------------------------------------------------------------------
# 测试环境
# ---------------------------------------------------------------------------


@pytest.fixture(scope="session")
def engine() -> Iterator[Engine]:
    """会话级 SQLite 内存引擎(StaticPool: 所有会话共享同一连接)。"""
    eng = create_engine(
        "sqlite+pysqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(eng)
    yield eng
    eng.dispose()


@pytest.fixture(autouse=True)
def _clean_state(engine: Engine) -> Iterator[None]:
    """每个测试前重置内存队列, 结束后清空全部表(避免测试间串扰)。"""
    queue.force_memory()
    yield
    with engine.begin() as conn:
        for table in reversed(Base.metadata.sorted_tables):
            conn.execute(table.delete())


@pytest.fixture()
def db(engine: Engine, tmp_path: Path) -> Iterator[Session]:
    """函数级共享会话(服务与测试共用, 端点内 commit); 对象存储指向临时目录。"""
    settings.data_dir = tmp_path
    factory = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    with factory() as session:
        yield session


@pytest.fixture()
def client(engine: Engine, db: Session, tmp_path: Path) -> Iterator[TestClient]:
    """测试客户端: 挂载项目/导出/管理路由, 替换 get_db 依赖。"""
    settings.data_dir = tmp_path
    app = create_app()
    app.include_router(projects_api.router)
    app.include_router(exports_api.router)
    app.include_router(admin_api.router)

    def _override_get_db() -> Iterator[Session]:
        yield db

    app.dependency_overrides[get_db] = _override_get_db
    with TestClient(app, raise_server_exceptions=False) as test_client:
        yield test_client


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------


def _h(client: TestClient, user) -> dict[str, str]:
    """认证头: 以真实窗口会话登录(同一 client 内缓存, 避免多窗口接管)。"""
    return login_headers(client, user)








def _create_project(client: TestClient, user, name: str = "导出测试项目") -> int:
    """创建项目并返回项目 id。"""
    resp = client.post("/api/projects", json={"name": name}, headers=_h(client, user))
    assert resp.status_code == 201, resp.text
    return resp.json()["project"]["id"]


def _create_version(client: TestClient, user, project_id: int) -> int:
    """创建项目版本并返回版本 id。"""
    resp = client.post(
        f"/api/projects/{project_id}/versions",
        json={"name": "基准版本 v1", "reason": "manual_save"},
        headers=_h(client, user),
    )
    assert resp.status_code == 201, resp.text
    return resp.json()["version"]["id"]


def _seed_dataset_version(db: Session, project_id: int, tag: str = "ds") -> int:
    """直接建数据集版本 + 内容寻址对象 + 文件行, 返回 dataset_version_id。"""
    user = db.execute(select(User).order_by(User.id)).scalars().first()
    assert user is not None
    csv_content = (
        b"timestamp,e_load,h_load,c_load\n"
        b"2025-01-01 00:00:00,125.5,85.2,60.0\n"
        b"2025-01-01 01:00:00,120.0,80.0,55.0\n"
    )
    obj = put_object(db, csv_content, "text/csv; charset=utf-8",
                                     source_category="dataset")
    dataset = Dataset(project_id=project_id, name=f"{tag}_数据集", status="published",
                      created_by=user.id)
    db.add(dataset)
    db.flush()
    version = DatasetVersion(
        dataset_id=dataset.id, version_no=1, timeline="hourly", resolution="1h",
        fixed_utc_offset_minutes=480, fields={"e_load": {"unit": "kWh"}},
        units={"e_load": "kWh"},
        quality_report={"checks": {"row_count": {"ok": True}}},
        provenance={"source_category": "test", "tag": tag},
        license="CC-BY-4.0",
        content_hash=obj.sha256, created_by=user.id, created_reason="test",
    )
    db.add(version)
    db.flush()
    db.add(DatasetFile(dataset_version_id=version.id, object_id=obj.id, file_kind="data",
                       format="csv", row_count=2, size_bytes=len(csv_content)))
    db.commit()
    return version.id


def _seed_evidence(
    db: Session, project_id: int, version_id: int, dataset_version_id: int, owner_id: int,
) -> tuple[int, int]:
    """直接建任务/快照/证据包/评估/结果索引, 返回 (evidence_package_id, assessment_id)。

    证据对象内容为 JSON: kpis/财务/环境/工程/适用(Excel 报告的摘要数据来源)。
    """
    evidence_content = json.dumps(
        {
            "kpis": [
                {"name": "年购电量", "value": 123456.7, "unit": "kWh"},
                {"name": "系统效率", "value": 92.5, "unit": "%"},
            ],
            "financial": {"总投资": 1000, "IRR": 0.12, "回收期": 8.5},
            "environmental": {"年碳排放": 123.4, "单位": "tCO2"},
            "engineering": {"装机容量": 500, "单位": "kW"},
            "applicability": {"适用范围": "华东地区商业综合体", "限制": "分时电价假设固定"},
        },
        ensure_ascii=False,
    ).encode("utf-8")
    obj = put_object(db, evidence_content, "application/json",
                                     source_category="evidence")
    snapshot = CalcSnapshot(
        project_version_id=version_id, dataset_version_ids=[dataset_version_id],
        calc_config_snapshot={
            "params": {"horizon_years": 1},
            "variables": [{"name": "pv_capacity"}],
            "objectives": [{"name": "min_cost"}],
            "constraints": [],
            "algorithm": "milp",
            "solver": "HiGHS",
            "tolerances": {"gap": 0.01},
            "random_seed": 42,
        },
        program_version="0.1.0", extension_versions={}, random_seed=42,
        tolerances={"gap": 0.01}, content_hash=sha256_hex(evidence_content),
        created_by=owner_id,
    )
    db.add(snapshot)
    db.flush()
    task = Task(project_id=project_id, type="calc", status="completed",
                business_outcome="normal_completion", calc_snapshot_id=snapshot.id,
                requested_by=owner_id)
    db.add(task)
    db.flush()
    package = EvidencePackage(
        task_id=task.id, calc_snapshot_id=snapshot.id, object_id=obj.id,
        content_hash=sha256_hex(evidence_content), status="complete", created_by=owner_id,
    )
    db.add(package)
    db.flush()
    assessment = ResultAssessment(
        evidence_package_id=package.id, assessor="system", assessed_by=owner_id,
        dimension_physical="pass", dimension_optimality="pass",
        dimension_financial="pass", dimension_reliability="pass",
        overall_score=88.5, comment="指标满足要求", detail={},
    )
    db.add(assessment)
    db.flush()
    db.add(
        ResultIndex(
            project_id=project_id, project_version_id=version_id,
            evidence_package_id=package.id, assessment_id=assessment.id,
            result_hash=sha256_hex(evidence_content), is_latest=True,
        )
    )
    db.commit()
    return package.id, assessment.id


def _download(
    client: TestClient, url: str, token: str, headers: dict | None = None
) -> tuple[int, bytes, str]:
    """带 token 下载(下载端点要求登录, 需传授权头), 返回 (status, content, content_type)。"""
    resp = client.get(url, params={"token": token}, headers=headers)
    return resp.status_code, resp.content, resp.headers.get("content-type", "")


def _expired_token(object_id: int, kind: str = "package") -> str:
    """构造已过期(exp 在过去)的签名下载授权(复刻 HMAC 签名方案)。"""
    payload = base64.urlsafe_b64encode(
        json.dumps(
            {"object_id": object_id, "kind": kind, "exp": int(time.time()) - 60},
            sort_keys=True,
        ).encode("utf-8")
    ).decode("ascii")
    sig = hmac.new(settings.secret_key.encode("utf-8"), payload.encode("utf-8"),
                   hashlib.sha256).hexdigest()
    return f"{payload}.{sig}"


def _tamper_package(zip_bytes: bytes) -> bytes:
    """重建 zip 并篡改 draft.json 内容(追加空格: 内容变化 → 校验值/大小必然不符)。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zin, zipfile.ZipFile(
        buf, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "draft.json":
                data = data + b" "  # 篡改: 与对象清单登记的 sha256/大小不一致
            zout.writestr(info.filename, data)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# 项目包导出(U15)
# ---------------------------------------------------------------------------


def test_export_package_owner_only_with_manifest_and_checksum(
    client: TestClient, db: Session,
) -> None:
    """仅所有者可导出项目包; 包含版本化清单/对象清单/逐对象内容校验。"""
    owner = make_user(db, "owner")
    viewer = make_user(db, "viewer")
    stranger = make_user(db, "stranger")
    pid = _create_project(client, owner)
    vid = _create_version(client, owner, pid)
    dvid = _seed_dataset_version(db, pid, tag="pkg")
    ep_id, a_id = _seed_evidence(db, pid, vid, dvid, owner.id)

    # 查看者 / 非成员 → 403
    for user in (viewer, stranger):
        resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, user))
        assert resp.status_code == 403, resp.text

    # 所有者导出 → 下载授权 + 清单
    resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, owner))
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["token"]
    assert body["manifest"]["format_version"] == "1.0"
    assert body["file_name"].endswith(".zip")

    status, content, ctype = _download(
        client, f"/api/projects/{pid}/exports/package/download", body["token"],
        headers=_h(client, owner),
    )
    assert status == 200
    assert "zip" in ctype
    assert content[:2] == b"PK"

    # 清单与逐对象校验值
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        manifest = json.loads(zf.read("manifest.json").decode("utf-8"))
        assert manifest["format_version"] == "1.0"
        assert manifest["package_type"] == "project"
        assert "project.json" in zf.namelist()
        assert "draft.json" in zf.namelist()
        objects_manifest = manifest["objects"]
        assert len(objects_manifest) >= 4  # 项目/草稿/数据集/证据
        for entry in objects_manifest:
            assert entry["path"] in zf.namelist()
            raw = zf.read(entry["path"])
            assert len(raw) == entry["size_bytes"]
            assert sha256_hex(raw) == entry["sha256"]
        # 不含账号/权限/会话/全局配置/密钥
        assert not ({"accounts", "permissions", "sessions", "global_config", "secrets"} & set(manifest))
        # 证据与评估引用随包保留
        assert any(p.startswith("evidence/") for p in zf.namelist())

    # 审计: 导出事件
    audit = db.execute(select(AuditLog).order_by(AuditLog.id.desc())).scalars().first()
    assert audit is not None and audit.action == "project.exported"
    assert audit.after["result"]["kind"] == "package"


def test_owner_can_export_excel_bilingual_and_fixed_reference(
    client: TestClient, db: Session,
) -> None:
    """所有者可导出 Excel(标题中英双语, 固定引用证据包与评估, 不重新求解)。"""
    owner = make_user(db, "owner")
    pid = _create_project(client, owner)
    vid = _create_version(client, owner, pid)
    dvid = _seed_dataset_version(db, pid, tag="xls")
    ep_id, a_id = _seed_evidence(db, pid, vid, dvid, owner.id)
    resp = client.post(
        f"/api/projects/{pid}/exports/excel",
        json={"evidence_package_id": ep_id, "assessment_id": a_id, "lang": "zh"},
        headers=_h(client, owner),
    )
    assert resp.status_code == 200, resp.text
    token = resp.json()["token"]

    status, content, ctype = _download(
        client, f"/api/projects/{pid}/exports/excel/download", token, headers=_h(client, owner)
    )
    assert status == 200
    assert "spreadsheetml" in ctype

    wb = load_workbook(io.BytesIO(content))
    ws = wb["报告总览"]
    title = ws.cell(row=1, column=1).value
    assert "pIES 项目结果报告" in title  # 中文标题
    assert "Project Result Report" in title  # 英文标题
    # 固定引用证据包与评估
    joined = "\n".join(
        str(ws.cell(row=r, column=1).value or "") + " " + str(ws.cell(row=r, column=2).value or "")
        for r in range(3, 24)
    )
    assert f"evidence_package={ep_id}" in joined
    assert f"assessment={a_id}" in joined
    assert "四维结论" in joined
    sheet_names = wb.sheetnames
    assert any("主要指标" in name for name in sheet_names)
    assert any("设备配置" in name for name in sheet_names)
    assert any("财务摘要" in name for name in sheet_names)
    assert any("环境摘要" in name for name in sheet_names)
    assert any("工程摘要" in name for name in sheet_names)
    # 主要指标表含证据包 KPI
    kp = wb[[n for n in sheet_names if "主要指标" in n][0]]
    kp_values = [str(kp.cell(row=r, column=1).value) for r in range(3, 8)]
    assert "年购电量" in kp_values


def test_non_owner_cannot_export_package_and_bad_identity(
    client: TestClient, db: Session,
) -> None:
    """项目不共享：非所有者不可导出项目包；缺认证头 401。"""
    owner = make_user(db, "owner")
    other_user = make_user(db, "other_user")
    pid = _create_project(client, owner)
    resp = client.post(
        f"/api/projects/{pid}/exports/package", headers=_h(client, other_user),
    )
    assert resp.status_code == 403, resp.text
    # 匿名(清空 cookie jar, 避免携带先前登录会话 Cookie)→ 401
    client.cookies.clear()
    resp = client.post(f"/api/projects/{pid}/exports/package")
    assert resp.status_code == 401, resp.text


# ---------------------------------------------------------------------------
# 项目包导入(U14)
# ---------------------------------------------------------------------------


def test_import_creates_new_identity_owner_and_evidence_source(
    client: TestClient, db: Session,
) -> None:
    """导入 → 新项目身份(不覆盖)/所有者=导入者/历史结果作为证据来源(不伪造任务)。"""
    owner = make_user(db, "owner")
    importer = make_user(db, "importer")
    pid = _create_project(client, owner)
    vid = _create_version(client, owner, pid)
    dvid = _seed_dataset_version(db, pid, tag="imp")
    ep_id, a_id = _seed_evidence(db, pid, vid, dvid, owner.id)

    resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, owner))
    zip_bytes = client.get(
        f"/api/projects/{pid}/exports/package/download",
        params={"token": resp.json()["token"]},
        headers=_h(client, owner),
    ).content
    original_evidence_count = len(
        db.execute(select(ObjectRef).where(ObjectRef.ref_type == "imported_evidence")).scalars().all()
    )

    proposal = package_service.import_proposal(
        db, importer, zip_bytes, idempotency_key="idem-import-1"
    )
    assert proposal.status == "proposed"
    assert proposal.review_summary["checks"]["integrity_ok"] is True
    # 幂等: 相同源文件重复提案返回同一提案
    again = package_service.import_proposal(db, importer, zip_bytes, idempotency_key="idem-import-1")
    assert again.id == proposal.id

    new_project = package_service.confirm_import(db, importer, proposal.id)
    db.commit()

    # 新项目身份: 与源项目不同, 名称不覆盖(自动去重后缀)
    old_project = db.get(Project, pid)
    assert new_project.id != pid
    assert new_project.owner_id == importer.id
    assert old_project.owner_id == owner.id
    assert new_project.name != old_project.name
    assert new_project.name.startswith("导出测试项目")

    # 原授权关系不迁移: 0.8.0 起无成员表, 导入者即新项目唯一所有者
    # (projects.owner_id == importer.id, 见上); 项目包内不携带账号权限。

    # 草稿内容迁移(模型/配置)且修订从 1 开始
    draft_content = project_service.get_current_draft_content(db, new_project.id)
    assert draft_content["calc_config"]["algorithm"] == "milp" or draft_content["model"]["devices"] == []

    # 历史结果作为证据来源保留(不伪造本地任务)
    assert len(draft_content.get("imported_evidence", [])) == 1
    assert draft_content["imported_evidence"][0]["package_id"] == ep_id
    assert draft_content["imported_evidence"][0]["task"]["business_outcome"] == "normal_completion"
    new_tasks = db.execute(select(Task).where(Task.project_id == new_project.id)).scalars().all()
    assert len(new_tasks) == 0  # 不伪造本地任务
    new_refs = db.execute(
        select(ObjectRef).where(ObjectRef.ref_type == "imported_evidence")
    ).scalars().all()
    assert len(new_refs) > original_evidence_count

    # 提案收尾 + 审计
    assert proposal.status == "applied"
    assert proposal.decided_by == importer.id
    # 幂等重放: 已导入提案再次确认返回同一项目
    same = package_service.confirm_import(db, importer, proposal.id)
    assert same.id == new_project.id

    # 非提案人确认 → 403
    other = make_user(db, "other")
    with pytest.raises(ForbiddenError):
        package_service.confirm_import(db, other, proposal.id)


def test_import_rejects_corrupt_and_non_zip(client: TestClient, db: Session) -> None:
    """导入校验失败拒绝: 篡改内容(完整性)与非 zip(格式)。"""
    owner = make_user(db, "owner")
    importer = make_user(db, "importer")
    pid = _create_project(client, owner)
    resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, owner))
    zip_bytes = client.get(
        f"/api/projects/{pid}/exports/package/download",
        params={"token": resp.json()["token"]},
        headers=_h(client, owner),
    ).content

    # 非 zip → 格式校验失败
    with pytest.raises(package_service.ImportValidationError):
        package_service.import_proposal(db, importer, b"not a zip at all")

    # 篡改对象内容 → 完整性校验失败(sha256/大小不符)
    tampered = _tamper_package(zip_bytes)
    with pytest.raises(package_service.ImportValidationError) as exc:
        package_service.import_proposal(db, importer, tampered)
    joined = "; ".join(exc.value.reasons)
    assert "校验值不符" in joined or "大小不符" in joined

    # 校验失败不创建任何提案/项目
    proposals = db.execute(select(ImportProposal)).scalars().all()
    assert proposals == []


def test_import_proposal_rejects_forbidden_sections(client: TestClient, db: Session) -> None:
    """包含账号/权限/会话等禁止内容的包拒绝导入(RPD 6)。"""
    owner = make_user(db, "owner")
    importer = make_user(db, "importer")
    pid = _create_project(client, owner)
    resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, owner))
    zip_bytes = client.get(
        f"/api/projects/{pid}/exports/package/download",
        params={"token": resp.json()["token"]},
        headers=_h(client, owner),
    ).content
    # 向清单注入禁止键
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zin, zipfile.ZipFile(
        buf, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == "manifest.json":
                manifest = json.loads(data.decode("utf-8"))
                manifest["permissions"] = {"owner": 1}
                data = json.dumps(manifest, ensure_ascii=False).encode("utf-8")
            zout.writestr(info.filename, data)
    with pytest.raises(package_service.ImportValidationError) as exc:
        package_service.import_proposal(db, importer, buf.getvalue())
    assert any("禁止内容" in r for r in exc.value.reasons)


# ---------------------------------------------------------------------------
# 下载授权(短期单对象)
# ---------------------------------------------------------------------------


def test_download_token_expired_and_tampered(client: TestClient, db: Session) -> None:
    """短期授权过期/伪造拒绝(下载 400)。"""
    owner = make_user(db, "owner")
    pid = _create_project(client, owner)
    resp = client.post(f"/api/projects/{pid}/exports/package", headers=_h(client, owner))
    object_id = resp.json()["object_id"]

    # 过期 token(exp 在过去)→ 服务层拒绝
    expired = _expired_token(object_id)
    with pytest.raises(package_service.DownloadTokenError):
        package_service.verify_download_token(expired, expected_kind="package")
    status, _content, _ct = _download(
        client, f"/api/projects/{pid}/exports/package/download", expired,
        headers=_h(client, owner),
    )
    assert status == 400

    # 篡改签名 → 拒绝
    valid = resp.json()["token"]
    tampered_token = valid[:-1] + ("0" if valid[-1] != "0" else "1")
    with pytest.raises(package_service.DownloadTokenError):
        package_service.verify_download_token(tampered_token, expected_kind="package")
    status, _content, _ct = _download(
        client, f"/api/projects/{pid}/exports/package/download", tampered_token,
        headers=_h(client, owner),
    )
    assert status == 400

    # 类型不符(excel token 用于包下载)→ 拒绝
    excel_token = package_service.create_download_token(
        object_id, "excel", project_id=pid, user_id=owner.id
    )
    with pytest.raises(package_service.DownloadTokenError):
        package_service.verify_download_token(excel_token, expected_kind="package")

    # 正常 token 校验通过(5 分钟窗口内)
    info = package_service.verify_download_token(valid, expected_kind="package")
    assert info["object_id"] == object_id
    assert info["project_id"] == pid and info["user_id"] == owner.id
    assert info["object_id"] == object_id


def test_download_rejects_object_not_in_project(client: TestClient, db: Session) -> None:
    """归属校验回归(0.2.0 A3 补丁): object 不属于该 project 时下载被拒。

    场景: 项目 A 导出对象 O; 攻击者(或普通用户)用「项目 B 的 URL + 自签
    project_id=B 的 token」请求下载。签名/绑定层会通过(user_id 是自己、
    project_id 一致), 但 object O 从未被项目 B 引用 → 归属校验拒绝。
    """
    owner = make_user(db, "owner")
    pid_a = _create_project(client, owner, name="归属校验项目A")
    pid_b = _create_project(client, owner, name="归属校验项目B")

    # 项目 A 导出项目包 → 拿到 object_id + 合法 token
    resp = client.post(f"/api/projects/{pid_a}/exports/package", headers=_h(client, owner))
    assert resp.status_code == 200, resp.text
    body = resp.json()
    object_id = body["object_id"]
    token_a = body["token"]

    # 项目 A 正常下载 → 200(该 object 被项目 A 引用)
    status, content, _ct = _download(
        client, f"/api/projects/{pid_a}/exports/package/download", token_a,
        headers=_h(client, owner),
    )
    assert status == 200, f"项目 A 下载应成功, got {status}"

    # 用「项目 B 的 URL + 自签 project_id=B 的 token」请求下载项目 A 的对象 → 400
    forged = package_service.create_download_token(
        object_id, "package", project_id=pid_b, user_id=owner.id
    )
    status, _content, _ct = _download(
        client, f"/api/projects/{pid_b}/exports/package/download", forged,
        headers=_h(client, owner),
    )
    assert status == 400, f"跨项目下载应被归属校验拒绝, got {status}"
    # 用项目 A 的 URL + 该伪造 token(project_id 不匹配)→ 同样 400(绑定层先拦)
    status, _content, _ct = _download(
        client, f"/api/projects/{pid_a}/exports/package/download", forged,
        headers=_h(client, owner),
    )
    assert status == 400, f"绑定层应拦截, got {status}"


# ---------------------------------------------------------------------------
# 管理端点(U16)
# ---------------------------------------------------------------------------


def test_admin_health_storage_and_audit(client: TestClient, db: Session) -> None:
    """管理员可读健康/存储/审计; 非管理员 403。"""
    admin = make_user(db, "admin1", role="admin")
    owner = make_user(db, "owner")
    pid = _create_project(client, owner)

    # 非管理员 → 403
    resp = client.get("/api/admin/health", headers=_h(client, owner))
    assert resp.status_code == 403, resp.text

    resp = client.get("/api/admin/health", headers=_h(client, admin))
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["status"] == "ok"
    assert body["liveness"]["ok"] is True
    assert body["readiness"]["db"] is True
    assert "queue" in body and "storage" in body and "metrics" in body

    resp = client.get("/api/admin/storage", headers=_h(client, admin))
    assert resp.status_code == 200
    # STO-07: 单一 StorageStatusDto(无兼容并集)
    body = resp.json()
    assert "objects" in body and "refs" in body and "capacity" in body
    assert "corrupt_count" in body and "cleanup_candidates" in body
    assert "stats" not in body and "sample_verify" not in body

    resp = client.get("/api/admin/audit", params={"entity_type": "project"}, headers=_h(client, admin))
    assert resp.status_code == 200
    items = resp.json()["items"]
    assert any(item["action"] == "project.created" and item["entity_id"] == pid for item in items)

    resp = client.get("/api/admin/audit", params={"action": "project.exported"}, headers=_h(client, admin))
    assert resp.status_code == 200 and resp.json()["items"] == []


def test_admin_unlock_task(client: TestClient, db: Session) -> None:
    """管理员解锁卡死任务: running → queued, 租约吊销, 槽释放, 全程审计。"""
    admin = make_user(db, "admin1", role="admin")
    owner = make_user(db, "owner")
    pid = _create_project(client, owner)
    now = datetime.now(UTC)
    task = Task(project_id=pid, type="calc", status="running", requested_by=owner.id)
    db.add(task)
    db.flush()
    attempt = TaskAttempt(task_id=task.id, attempt_no=1, worker_id="w1",
                          status="running", started_at=now)
    db.add(attempt)
    db.flush()
    db.add(TaskLease(attempt_id=attempt.id, lease_token=uuid4(), acquired_by="w1",
                     acquired_at=now, renewed_at=now,
                     expires_at=now + timedelta(seconds=60), status="active"))
    db.add(ComputeSlot(pool_name="compute", status="busy", capacity=1, in_use=1,
                       current_attempt_id=attempt.id))
    db.commit()

    # 0.2.0 B2: 未携带 confirm → 409(危险操作确认), 任务保持 running 不被改动
    resp = client.post(
        "/api/admin/unlock-task", json={"task_id": task.id}, headers=_h(client, admin)
    )
    assert resp.status_code == 409, resp.text
    assert resp.json()["error"]["code"] == "ADMIN-CONFIRM-REQUIRED"
    db.refresh(task)
    assert task.status == "running"
    db.refresh(attempt)
    assert attempt.status == "running"

    # 携带 confirm=true → 200, 解锁执行
    resp = client.post(
        "/api/admin/unlock-task", json={"task_id": task.id, "confirm": True},
        headers=_h(client, admin),
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["unlocked"] is True
    db.refresh(task)
    assert task.status == "queued"
    db.refresh(attempt)
    assert attempt.status == "stopped" and attempt.stop_reason == "admin_unlock"
    lease = db.execute(select(TaskLease).where(TaskLease.attempt_id == attempt.id)).scalar_one()
    assert lease.status == "revoked"
    slot = db.execute(select(ComputeSlot).where(ComputeSlot.pool_name == "compute")).scalar_one()
    assert slot.in_use == 0 and slot.current_attempt_id is None

    # 审计 + 维护记录
    audit = db.execute(
        select(AuditLog).where(AuditLog.action == "maintenance.unlock_task")
    ).scalar_one()
    assert audit.actor_id == admin.id and audit.actor_type == "admin"
    maintenance = db.execute(select(AdminMaintenanceAction)).scalar_one()
    assert maintenance.action_type == "user_override"

    # 终态任务不可解锁
    task.status = "completed"
    db.commit()
    resp = client.post(
        "/api/admin/unlock-task", json={"task_id": task.id, "confirm": True},
        headers=_h(client, admin),
    )
    assert resp.status_code == 409
